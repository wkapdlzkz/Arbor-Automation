import requests
import re
import openpyxl
from bs4 import BeautifulSoup
import urllib3
from pprint import pprint
from urllib import parse
from selenium import webdriver
from datetime import datetime
import time
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import Rule
from openpyxl.formatting.rule import CellIsRule
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

#SSL 인증서 검증 OFF 에러 예외처리
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


#세션만들기

# 엑셀 저장

##############################################################################################################
#설정값

# 1000개 라인 호출 변수
linenum = 3000

start_range = 1234        #시작 ID (ALL Mitigation URL ID 참고)    Set index id in url query part of the Arbor all Mitigation tab to match the period you want to look up
xxend_range = 1234+1      #끝나는 ID (ALL Mitigation URL ID 참고)  Set index id in url query part of the Arbor all Mitigation tab to match the period you want to look up

start_month = 1
stop_month = 2


##############################################################################################################

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'sheet'
# 컬럼명 지정(헤더)

sheet['A1'] = '일자'
sheet['B1'] = 'Alert ID'
sheet['C1'] = '고객사MO명'
sheet['D1'] = '목적지IP'
sheet['E1'] = 'CP 탐지시간'
sheet['F1'] = 'CP 탐지명'
sheet['G1'] = 'CP 탐지량'
sheet['H1'] = 'TMS 이벤트'
sheet['I1'] = 'TMS 차단 시간'
sheet['J1'] = 'TMS 최대 유입량'
sheet['K1'] = 'TMS 평균 유입량'
sheet['L1'] = 'TMS 평균 차단률'
sheet['M1'] = '검증'



options = webdriver.ChromeOptions()

options.add_argument('--ignore-certificate-errors')
options.add_argument('--ignore-ssl-errors')
options.add_argument('--headless')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--disable-extensions')

# 드라이버 위치 경로 입력
driver = webdriver.Chrome("path to chrome driver", chrome_options=options)

# url을 이용하여 브라우저로 접속
driver.get('https://1.2.3.4/index')
driver.implicitly_wait(5)

driver.find_element_by_xpath('/html/body/div/form/label[1]/input').send_keys('id')
driver.find_element_by_xpath('/html/body/div/form/label[2]/input').send_keys('pw')

driver.find_element_by_xpath('/html/body/div/form/button').click()

_cookies = driver.get_cookies()
cookie_dict = {}
for cookie in _cookies:
    cookie_dict[cookie['name']] = cookie['value']


count  = 2

for i in range(start_range, xxend_range):
    url = f'https://1.2.3.4/page?id=mitigation_status&mitigation_id={i}' #Creating an Event ID
    driver.get(url)
    test = driver.find_element_by_xpath('//*[@id="active_summary_graph_div"]/div').text

    EVTMSlist = driver.find_element_by_xpath('/html/body/div[1]/div[1]/form/div[1]/div/section[1]/div/div[1]/div[1]/div/div[3]/p/a').text
    TMTMSlist = driver.find_element_by_xpath('/html/body/div[1]/div[1]/form/div[1]/div/section[1]/div/div[1]/div[1]/div/div[1]/p').text
    AVTMSlist = driver.find_element_by_xpath('/html/body/div[1]/div[1]/form/div[1]/div/section[1]/div/div[2]/div[2]/div[3]/table/tbody/tr[4]/td[4]').text
    MTTMSlist = driver.find_element_by_xpath('/html/body/div[1]/div[1]/form/div[1]/div/section[1]/div/div[2]/div[2]/div[3]/table/tbody/tr[5]/td[4]').text

    sheet[f'H{count}'] = EVTMSlist
    sheet[f'I{count}'] = TMTMSlist
    sheet[f'K{count}'] = AVTMSlist
    sheet[f'L{count}'] = MTTMSlist

    print(EVTMSlist)
    print(TMTMSlist)
    print(AVTMSlist)
    print(MTTMSlist)

    #bps
    test_check = re.findall("G|M|k",test)
    if(len(test_check) == 0):
       test_bps = re.sub(".*\n","",test)
       MXTMSlist = ("약 " + test_bps + " bps")
       sheet[f'J{count}'] = MXTMSlist
       print(MXTMSlist)
       count += 1

    else:
        #G M k bps
        test_1 = re.sub("G"," G",test)
        test_2 = re.sub("M"," M",test_1)
        test_3 = re.sub("k"," k",test_2)
        test_4 = re.sub(".*\n","",test_3)
        MXTMSlist = ("약 " + test_4 + "bps")
        sheet[f'J{count}'] = MXTMSlist
        count += 1
        print(MXTMSlist)

        time.sleep(1)

driver.quit()

session=requests.session()

#로그인 하는 페이지의 general-requestURL에서 url 가져옴
url="https://1.2.3.4/index"

data={
    "username": "id", 
    "password": "pw", 
    "Submit": "Log In"
}

response=session.post(url, data=data, verify=False)

url = 'https://1.2.3.4/page?id=my_sightline'
response=session.get(url, verify=False, allow_redirects=True)

CSRFName = re.findall("CSRFName.*\">",response.text)            
CSRFName = re.sub("\"|.*value=|>", "", CSRFName[0])

CSRFToken = re.findall("CSRFToken.*\">",response.text)
CSRFToken = re.sub("\"|.*value=|>", "", CSRFToken[0])

url = 'https://1.2.3.4/page?id=alerts_all&rpc=sort'
data = 'id=alerts_all&rpc=sort'
response=session.post(url, data=data, verify=False, allow_redirects=True)



cache_id = response.url

url = cache_id


data={
"id": "alerts_all",
"cache_id": cache_id,
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_search_string": "ac:\"DoS\" sev:high",
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_search_button": "Search",
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_use_start": "on",
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_start_dir": "after",
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_start_month1": start_month,
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_start_day1": "1",
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_start_year1": "2023",
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_start_hour1": "00:00",
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_use_stop": "on",
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_stop_dir": "before",
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_stop_month1": stop_month,
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_stop_day1": "1",
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_stop_year1": "2023",
"AlertSearchWidget_cb602aee0baafe0bb676a818d887bd77_stop_hour1": "00:00",
"AlertListingWidget_f236dd0cbc80e5a8eae3dfab537f7639_nav_sort_field": "AlertListingWidget_f236dd0cbc80e5a8eae3dfab537f7639_alert_id_col",
"AlertListingWidget_f236dd0cbc80e5a8eae3dfab537f7639_nav_sort_direction": "descending",
"AlertListingWidget_f236dd0cbc80e5a8eae3dfab537f7639_currpage": "1",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_alert_class_saved": "dos",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_alert_type_saved": "dos_all",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_start_wiz_dir_saved": "after",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_stop_wiz_dir_saved": "before",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_importance_high": "on",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_alert_class": "dos",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_alert_type": "dos_all",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_alert_classification": "all",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_search_limit": linenum,
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_page_size": linenum,
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_ongoing": "on",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_recent": "on",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_use_start_wiz": "on",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_start_wiz_dir": "after",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_start_wiz_month1": start_month,
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_start_wiz_day1": "1",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_start_wiz_year1": "2023",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_start_wiz_hour1": "00:00",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_use_stop_wiz": "on",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_stop_wiz_dir": "before",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_stop_wiz_month1": stop_month,
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_stop_wiz_day1": "1",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_stop_wiz_year1": "2023",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_stop_wiz_hour1": "00:00",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_lo_bps_wiz_base": "",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_lo_bps_wiz_scale": "u",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_hi_bps_wiz_base": "",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_hi_bps_wiz_scale": "u",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_lo_pps_wiz_base": "",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_lo_pps_wiz_scale": "u",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_hi_pps_wiz_base": "",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_hi_pps_wiz_scale": "u",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_low_sev_wiz": "",
"AlertSearchPopIn_63176063afc589a39b359784afa7c413_high_sev_wiz": "",
"CSRFName": CSRFName,
"CSRFToken": CSRFToken,
"id": "alerts_all",
"widget_id": "AlertListingWidget_f236dd0cbc80e5a8eae3dfab537f7639",
"sprpcv": "2",
"rpc": "sort",
"last_sort": "AlertListingWidget_f236dd0cbc80e5a8eae3dfab537f7639_alert_id_col",
"sort_field": "AlertListingWidget_f236dd0cbc80e5a8eae3dfab537f7639_alert_id_col"
}

response=session.post(url, data=data, verify=False)




result  = response.text
result = parse.unquote(result, encoding="utf-8")

# 본문 전체 코드 개행 분리

result = re.sub("</div>\\n</div>", "", result)
result = re.sub("<br/>\\n</div>", "\\n", result)
result = re.sub("\\t\\n</div>", "", result)
result = re.sub("\\n</div>", "-#-#-\\n", result)
result_txt = re.findall(".*", result)

count  = 2

for start in result_txt:
    
    #IDlist
    if "<tr>    <td \t\t\t\t\tstyle=\"width:5%\"><a href=\"/page?id=host_alert&amp;alert_id=" in start:
        start = re.sub("<.*\">|<.*", "", start)
        IDlist = start
        continue

        
    #PFlist    

    if "using" in start:
        start = re.sub("</a>.*", "", start)
        start = re.sub("^.*strong\">", "", start)
        PFlist = start
        continue

    #IPlist


    if "Incoming Host Alert to  " in start:
        start = re.sub("^.*to  |<.*", "", start)
        IPlist = start
        continue
        
    #EVlist
    if "-#-#-" in start:
        start = re.sub("-#-#-", "", start)
        EVlist = start
        continue

    #TRlist    
    if "alert_importance_high" in start:
        start = start.split("<br>", 2)
        start = re.sub("<br>.*", "", start[2])
        start = re.sub("&nbsp;", " ", start)
        TRlist = start
        continue

        
    #TIlist
    if "<span class=\"alert_duration\">" in start:
 
        if "Ongoing" in start:
            continue

        if "Global Detection" in PFlist:
            continue

        if "Outgoing Host" in IPlist:
            continue

        start = re.sub(".*alert_duration\">|</span>|</t.*", "", start)
        TIlist = start
        date = re.sub(" [0-9][0-9]:.*", "", start)
        date = datetime.datetime.strptime(date, '%b %d').strftime('2023%m%d')
        
        sheet[f'A{count}'] = date
        sheet[f'B{count}'] = IDlist
        sheet[f'C{count}'] = PFlist
        sheet[f'D{count}'] = IPlist
        sheet[f'E{count}'] = TIlist
        sheet[f'F{count}'] = EVlist
        sheet[f'G{count}'] = TRlist
        count += 1

wb.save('path.. to.. /ddoslist.xlsx')


# 엑셀 파일 불러오기
wb = openpyxl.load_workbook('path.. to.. /ddoslist.xlsx')

# 작업할 시트 선택하기
sheet = wb['sheet']

# 두 번째 행부터 시작하여 마지막 행까지 루프를 돌면서 비교
for row in range(2, sheet.max_row + 1):
    if sheet.cell(row=row, column=2).value == sheet.cell(row=row, column=8).value:
        sheet.cell(row=row, column=13).value = "good"
    else:
        sheet.cell(row=row, column=13).value = "bad"

# 현재 날짜와 시간을 문자열로 변환
now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")

# 변경된 엑셀 파일 저장
wb.save(f"path.. to.. /ddoslist_{now}.xlsx")
wb.close()
exit()